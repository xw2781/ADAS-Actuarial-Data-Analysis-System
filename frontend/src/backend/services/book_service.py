"""Workbook/Excel file operations via openpyxl."""
from __future__ import annotations

import os
import json
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from fastapi import HTTPException

from backend import config


def resolve_allowed_book(path_str: str) -> Path:
    p = Path(path_str).resolve()
    for root in config.ALLOWED_BOOK_DIRS:
        if p == root or str(p).startswith(str(root) + os.sep):
            return p
    raise HTTPException(400, "Workbook path not allowed.")


def get_book_mtime(path: str) -> float:
    return os.stat(path).st_mtime


def read_sheet_matrix(path: str, sheet_name: str, max_rows: int = 200, max_cols: int = 50):
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        raise HTTPException(404, f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    max_r = min(ws.max_row or 1, max_rows)
    max_c = min(ws.max_column or 1, max_cols)
    rows = []
    for r in range(1, max_r + 1):
        row = []
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            row.append(v)
        rows.append(row)
    return rows


def _normalize_matrix_rows(rows: List[List[Any]], max_rows: int = 200, max_cols: int = 50) -> List[List[Any]]:
    if not rows:
        return []
    limited = rows[:max_rows]
    width = min(max((len(r) for r in limited), default=0), max_cols)
    if width <= 0:
        return []
    out: List[List[Any]] = []
    for r in limited:
        row = list(r[:width])
        if len(row) < width:
            row.extend([None] * (width - len(row)))
        out.append(row)
    return out


def _load_project_map_data(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, dict):
        raise HTTPException(400, "Invalid project map format.")
    return data


def _project_map_sheet_names(data: Dict[str, Any]) -> List[str]:
    out: List[str] = []
    for k, v in data.items():
        if isinstance(v, dict) and isinstance(v.get("headers"), list) and isinstance(v.get("rows"), list):
            out.append(k)
    return out


def _read_project_map_sheet_matrix(
    path: str, sheet_name: str, max_rows: int = 200, max_cols: int = 50
) -> List[List[Any]]:
    data = _load_project_map_data(path)
    sheet_names = _project_map_sheet_names(data)
    if sheet_name not in sheet_names:
        raise HTTPException(404, f"Sheet not found: {sheet_name}")

    sheet = data.get(sheet_name) or {}
    headers = sheet.get("headers") or []
    body_rows = sheet.get("rows") or []
    rows: List[List[Any]] = [list(headers)]
    for r in body_rows:
        rows.append(list(r) if isinstance(r, list) else [])
    return _normalize_matrix_rows(rows, max_rows=max_rows, max_cols=max_cols)
