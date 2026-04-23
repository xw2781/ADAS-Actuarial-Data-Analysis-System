"""Excel COM interop operations (win32com)."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

import openpyxl


def excel_active_selection() -> Dict[str, Any]:
    try:
        import pythoncom
        import win32com.client
    except ImportError:
        return {"ok": False, "error": "win32com is not available on this system."}
    try:
        pythoncom.CoInitialize()
        xl = win32com.client.GetObject(Class="Excel.Application")
        wb = xl.ActiveWorkbook
        if wb is None:
            return {"ok": False, "error": "No active workbook in Excel."}
        book_path = wb.FullName
        sheet = xl.ActiveSheet.Name
        cell_addr = xl.ActiveCell.Address.replace("$", "")
        cell_value = xl.ActiveCell.Value
        numeric = None
        if cell_value is not None:
            try:
                numeric = float(cell_value)
            except (ValueError, TypeError):
                return {"ok": False, "error": f"Cell value is not numeric: {repr(cell_value)}"}
        return {
            "ok": True,
            "book_path": book_path,
            "sheet": sheet,
            "cell": cell_addr,
            "value": numeric,
        }
    except Exception as e:
        return {"ok": False, "error": f"Cannot connect to Excel: {str(e)}"}


def excel_wait_for_enter() -> Dict[str, Any]:
    try:
        import pythoncom
        import win32com.client
        import win32api
        import win32gui
    except ImportError:
        return {"ok": False, "error": "win32com is not available on this system."}
    try:
        pythoncom.CoInitialize()
        xl = win32com.client.GetObject(Class="Excel.Application")
        wb = xl.ActiveWorkbook
        if wb is None:
            return {"ok": False, "error": "No active workbook in Excel."}
        excel_hwnd = xl.Hwnd
        import time
        VK_RETURN = 0x0D
        confirmed = False
        was_down = False
        last_addr = xl.ActiveCell.Address.replace("$", "")
        last_sheet = xl.ActiveSheet.Name
        last_book = wb.FullName
        last_value = xl.ActiveCell.Value
        for _ in range(300):
            pythoncom.PumpWaitingMessages()
            time.sleep(0.1)
            try:
                pythoncom.PumpWaitingMessages()
                cur_addr = xl.ActiveCell.Address.replace("$", "")
                cur_sheet = xl.ActiveSheet.Name
                cur_book = xl.ActiveWorkbook.FullName if xl.ActiveWorkbook else last_book
                cur_value = xl.ActiveCell.Value
                key_state = win32api.GetAsyncKeyState(VK_RETURN)
                is_down = (key_state & 0x8000) != 0
                fg_hwnd = win32gui.GetForegroundWindow()
                if is_down and not was_down and fg_hwnd == excel_hwnd:
                    confirmed = True
                    break
                was_down = is_down
                last_addr = cur_addr
                last_sheet = cur_sheet
                last_book = cur_book
                last_value = cur_value
            except Exception:
                break
        cell_addr = last_addr
        sheet_name = last_sheet
        book_path = last_book
        numeric = None
        if last_value is not None:
            try:
                numeric = float(last_value)
            except (ValueError, TypeError):
                pass
        return {
            "ok": True,
            "confirmed": confirmed,
            "book_path": book_path,
            "sheet": sheet_name,
            "cell": cell_addr,
            "value": numeric,
        }
    except Exception as e:
        return {"ok": False, "error": f"Cannot connect to Excel: {str(e)}"}


def excel_read_cell(book_path: str, sheet: str, cell: str) -> Dict[str, Any]:
    book = Path(book_path).resolve()
    if not book.exists():
        return {"ok": False, "error": f"File not found: {book_path}"}
    try:
        wb = openpyxl.load_workbook(str(book), data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return {"ok": False, "error": f"Sheet not found: {sheet}"}
        ws = wb[sheet]
        cell_value = ws[cell].value
        wb.close()
        numeric = None
        if cell_value is not None:
            try:
                numeric = float(cell_value)
            except (ValueError, TypeError):
                return {"ok": False, "error": f"Cell value is not numeric: {repr(cell_value)}"}
        return {"ok": True, "value": numeric}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def excel_read_cells_batch(items: list) -> Dict[str, Any]:
    from collections import defaultdict
    groups: Dict[str, list] = defaultdict(list)
    for i, item in enumerate(items):
        groups[item.book_path].append((i, item))

    results = [None] * len(items)
    for book_path, batch_items in groups.items():
        p = Path(book_path).resolve()
        if not p.exists():
            for i, item in batch_items:
                results[i] = {"ok": False, "error": f"File not found: {book_path}"}
            continue
        try:
            wb = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
            for i, item in batch_items:
                if item.sheet not in wb.sheetnames:
                    results[i] = {"ok": False, "error": f"Sheet not found: {item.sheet}"}
                    continue
                ws = wb[item.sheet]
                val = ws[item.cell].value
                try:
                    numeric = float(val) if val is not None else None
                    results[i] = {"ok": True, "value": numeric}
                except (ValueError, TypeError):
                    results[i] = {"ok": False, "error": f"Not numeric: {repr(val)}"}
            wb.close()
        except Exception as e:
            for i, item in batch_items:
                if results[i] is None:
                    results[i] = {"ok": False, "error": str(e)}

    return {"ok": True, "results": results}


def excel_open_workbook(book_path: str, sheet: str = "", cell: str = "") -> Dict[str, Any]:
    p = Path(book_path).resolve()
    if not p.exists():
        return {"ok": False, "error": f"File not found: {book_path}"}
    try:
        import pythoncom
        import win32com.client
    except ImportError:
        return {"ok": False, "error": "win32com is not available on this system."}
    try:
        pythoncom.CoInitialize()
        try:
            xl = win32com.client.GetObject(Class="Excel.Application")
        except Exception:
            xl = win32com.client.Dispatch("Excel.Application")
            xl.Visible = True
        full_path = str(p)
        target_wb = None
        already_open = False
        for wb in xl.Workbooks:
            if wb.FullName.lower() == full_path.lower():
                target_wb = wb
                already_open = True
                break
        if target_wb is None:
            target_wb = xl.Workbooks.Open(str(p), ReadOnly=True)
        xl.Visible = True
        try:
            import win32gui
            hwnd = xl.Hwnd
            import win32con
            if win32gui.IsIconic(hwnd):
                win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            win32gui.SetForegroundWindow(hwnd)
        except Exception:
            pass
        target_wb.Activate()
        if sheet and cell:
            try:
                ws = target_wb.Sheets(sheet)
                target = ws.Range(cell)
                scroll_row = max(1, target.Row - 10)
                scroll_col = max(1, target.Column - 10)
                ws.Activate()
                xl.Goto(ws.Cells(scroll_row, scroll_col), True)
                target.Select()
            except Exception:
                pass
        elif sheet:
            try:
                target_wb.Sheets(sheet).Activate()
            except Exception:
                pass
        return {"ok": True, "already_open": already_open}
    except Exception as e:
        return {"ok": False, "error": f"Failed to open workbook: {str(e)}"}
