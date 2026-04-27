from __future__ import annotations

import os
from typing import Any, Dict

import openpyxl
from fastapi import APIRouter, HTTPException

from backend.schemas.book import AnyBookSheetRequest, AnyBookPatchRequest
from backend.services import book_service

router = APIRouter()


@router.post("/book/meta")
def book_meta(req: AnyBookSheetRequest) -> Dict[str, Any]:
    book = book_service.resolve_allowed_book(req.book_path)
    if not book.exists():
        raise HTTPException(404, f"Workbook not found: {book}")
    st = os.stat(book)
    wb = openpyxl.load_workbook(str(book), read_only=True, keep_vba=True)
    return {"path": str(book), "mtime": st.st_mtime, "sheets": wb.sheetnames}


@router.post("/book/sheet")
def book_sheet(req: AnyBookSheetRequest) -> Dict[str, Any]:
    book = book_service.resolve_allowed_book(req.book_path)
    if not book.exists():
        raise HTTPException(404, f"Workbook not found: {book}")
    st = os.stat(book)
    values = book_service.read_sheet_matrix(str(book), sheet_name=req.sheet)
    return {"path": str(book), "sheet": req.sheet, "values": values, "mtime": st.st_mtime}


@router.post("/book/patch")
def book_patch(req: AnyBookPatchRequest) -> Dict[str, Any]:
    book = book_service.resolve_allowed_book(req.book_path)
    if not book.exists():
        raise HTTPException(404, f"Workbook not found: {book}")

    st = os.stat(book)
    if req.file_mtime is not None and abs(st.st_mtime - req.file_mtime) > 1e-6:
        raise HTTPException(409, "Workbook changed on disk. Reload and retry.")

    try:
        wb = openpyxl.load_workbook(str(book), data_only=True, keep_vba=True)
        if req.sheet not in wb.sheetnames:
            raise HTTPException(404, f"Sheet not found: {req.sheet}")
        ws = wb[req.sheet]

        for it in req.items:
            ws.cell(row=it.r + 1, column=it.c + 1).value = it.value

        tmp_path = str(book) + ".tmp"
        wb.save(tmp_path)
        os.replace(tmp_path, str(book))

        st2 = os.stat(book)
        return {"ok": True, "mtime": st2.st_mtime}

    except PermissionError:
        raise HTTPException(423, "Workbook is locked (possibly open in Excel). Close it and retry.")
